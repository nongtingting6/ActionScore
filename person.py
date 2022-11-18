import datetime
from openpyxl import load_workbook


class Person():

    def __init__(self, name):
        self.action_and_time = []
        self.name = name
        self.list, self.score = self.get_person_list("data.xlsx", name)

    def person_do_action(self, action_id):
        """
        当
        :param action_id:
        :return:
        """
        # 获取当前动作的信息，是列表[动作，范围，频率，分数]
        action_info_cur = self.list.get(action_id)

        # 如果当前动作id在excel的所有动作信息里的话
        if action_info_cur is not None:
            cur_score = action_info_cur[3]
            cur_action_info = action_info_cur[0]
        else:
            cur_score = 0
            cur_action_info = '输入错误'

        # 获取当前时间的datetime对象
        cur_day = datetime.datetime.now()
        # datatime格式化输出
        cur_time = cur_day.strftime('%Y年%m月%d日%H点%M分%S秒')

        # 累加保存当前对象的分数和动作记录
        self.score = self.score + cur_score
        self.action_and_time = self.action_and_time + [[cur_time,cur_action_info]]

    def get_person_list(self, excel_name, excel_sheet_name):
        """
        获取对象的信息
        :param excel_name: excel的名称
        :param excel_sheet_name: 选择操作的工作簿名称
        :return:第一个值:返回字典信息,对应为{'动作id':[动作描述,动作范围,动作频率,动作分数],'下一个动作id':[...]}
        第二个值:返回字典信息,对应为{'分数':当前工作簿分数统计}
        """
        date = load_workbook(excel_name)  # 加载
        sheet = date[excel_sheet_name]  # 选择编辑的工作簿
        action_to_info = {}
        score = float(sheet['B1'].value)

        # 1.按行循环工作表
        for row in sheet:
            row_value = []
            # 如果当前行第一个单元格的值是编号,则说明规则部分已经读取完成了
            if row[0].value == '编号':
                break
            # 2.循环每一行
            for cell in row:
                # 3.将每一行的单元格的值储存到列表
                row_value = row_value + [cell.value]
                # 4.一行循环完后,提取出要的数据形成字典
                action_to_info_temp = {str(row_value[0]): row_value[1:]}
                # 5.把字典添加到全局变量的字典里
                action_to_info.update(action_to_info_temp)

        # 6.把多余的key删除
        action_to_info.pop('分数')
        action_to_info.pop('索引')
        return action_to_info, score

    def write_record_to_excel(self):
        """
        将对象的属性写入到excel
        :return:
        """
        data = load_workbook('data.xlsx')
        sheet = data[self.name]
        end_row = sheet.max_row
        end_col = sheet.max_column
        # 保存做事的时间和事情记录
        for i in range(0,len(self.action_and_time)):
            last_index = sheet[f'A{end_row+i}'].value
            sheet[f'A{end_row+i +1}'] = 1+ int(last_index) if last_index !='编号' else 1
            sheet[f'B{end_row+i+1}'] = self.action_and_time[i][0]
            sheet[f'C{end_row+i+1}'] = self.action_and_time[i][1]
        self.action_and_time = []
        # 保存分数
        sheet['B1'] = self.score
        data.save('data.xlsx')
        data.close()